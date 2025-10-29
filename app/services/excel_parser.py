from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import colornames


def rgb_to_color_name(rgb_hex):
    """
    Convert RGB hex value to human-readable color name using colornames library.
    
    Args:
        rgb_hex (str): RGB hex value (e.g., "FF0000")
        
    Returns:
        str: Color name from colornames or hex value if not found
    """
    try:
        # Convert hex to RGB tuple
        rgb_tuple = tuple(int(rgb_hex[i:i+2], 16) for i in (0, 2, 4))
        
        # Get the nearest color name using colornames
        color_name = colornames.find(rgb_tuple)
        return color_name
                    
    except (ValueError, IndexError, Exception):
        # If parsing fails or color not found, return the hex value
        return f"#{rgb_hex}"


def get_cell_color_info(cell):
    """
    Extract color information from a cell.
    
    Args:
        cell: openpyxl cell object
        
    Returns:
        str: Color information in format "color:color_name" or None if no color
    """
    if cell.fill and cell.fill.fill_type == 'solid':
        if cell.fill.start_color and cell.fill.start_color.rgb:
            try:
                # Get RGB object and convert to hex string properly
                rgb_obj = cell.fill.start_color.rgb
                
                # Convert RGB object to hex string
                if hasattr(rgb_obj, 'rgb'):
                    rgb_hex = str(rgb_obj.rgb)
                elif hasattr(rgb_obj, 'value'):
                    rgb_hex = str(rgb_obj.value)
                else:
                    # Try different approaches to get the hex value
                    try:
                        # Try to get the hex representation directly
                        rgb_hex = rgb_obj.hex
                    except AttributeError:
                        # Fallback to string conversion
                        rgb_hex = str(rgb_obj)
                
                # Clean up the hex string
                if isinstance(rgb_hex, str) and rgb_hex:
                    # Remove any non-hex characters and ensure it's uppercase
                    rgb_hex = ''.join(c for c in rgb_hex.upper() if c in '0123456789ABCDEF')
                    
                    # Skip default/transparent colors
                    if rgb_hex in ['00000000', 'FFFFFFFF', '000000', 'FFFFFF']:
                        return None
                    
                    # Remove alpha channel if present (last 2 characters)
                    if len(rgb_hex) == 8:
                        rgb_hex = rgb_hex[2:]  # Remove alpha channel
                    
                    # Ensure we have a valid 6-character hex string
                    if len(rgb_hex) == 6:
                        # Convert to human-readable color name
                        color_name = rgb_to_color_name(rgb_hex)
                        return f"color:{color_name}"
                    
            except Exception as e:
                # If there's any error in color extraction, skip it silently
                return None
    return None


def excel_to_table_text(file_path: str, historical: bool = True) -> str:
    """
    Convert Excel file into a text representation.
    
    Args:
        file_path (str): Path to Excel file.
        historical (bool): 
            - True → only keep chosen values (ignore dropdown options), show colors for ALL cells
            - False → include dropdown options, show colors only for empty cells
    
    Returns:
        str: Text representation including:
            - Cell values and coordinates
            - Dropdown options (when historical=False)
            - Cell background colors (format: color:color_name)
            - Cell comments
            - Empty cells with formatting (colors) are included
    """
    workbook = load_workbook(file_path, data_only=True)
    output_lines = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet.sheet_state != "visible":
            continue  # skip hidden sheets
        output_lines.append(f"[Sheet: {sheet_name}]")

        # Build map of cell → dropdown formula
        dropdowns = {}
        if sheet.data_validations:
            for dv in sheet.data_validations.dataValidation:
                if dv.formula1:
                    for ref in dv.sqref:  # MultiCellRange
                        min_col, min_row, max_col, max_row = range_boundaries(str(ref))
                        for r in range(min_row, max_row + 1):
                            for c in range(min_col, max_col + 1):
                                coord = sheet.cell(row=r, column=c).coordinate
                                dropdowns[coord] = dv.formula1

        # Iterate rows
        for row in sheet.iter_rows(values_only=False):
            row_has_data = False
            row_lines = [f"Row {row[0].row}"]

            for cell in row:
                # Get color information first
                color_info = get_cell_color_info(cell)
                # print(color_info)
                
                # Only include cells that have content OR are empty but colored
                if cell.value is None and not cell.comment and not color_info:
                    continue
                row_has_data = True
                dropdown_options = None

                # Check if this cell has a dropdown
                if cell.coordinate in dropdowns and not historical:
                    formula = dropdowns[cell.coordinate]
                    if formula.startswith("{"):
                        # Inline list {"Yes","No","Maybe"}
                        dropdown_options = [opt.strip('"') for opt in formula.strip("{}").split(",")]
                    else:
                        # Cell range reference like A1:A3
                        try:
                            min_col, min_row, max_col, max_row = range_boundaries(formula)
                            dropdown_values = []
                            for r in sheet.iter_rows(min_row=min_row, max_row=max_row,
                                                     min_col=min_col, max_col=max_col):
                                for c in r:
                                    if c.value is not None:
                                        dropdown_values.append(str(c.value))
                            dropdown_options = dropdown_values
                        except Exception:
                            dropdown_options = [formula]

                # Build line
                line = f"{cell.coordinate} = \"{cell.value}\"" if cell.value is not None else f"{cell.coordinate} = (empty)"
                if dropdown_options:
                    line += f" (Options: {' | '.join(dropdown_options)})"
                # Show color based on historical mode
                if color_info:
                    if historical:
                        # Historical mode: show colors for all cells
                        line += f" ({color_info})"
                    else:
                        # Non-historical mode: only show colors for empty cells
                        if cell.value is None:
                            line += f" ({color_info})"
                if cell.comment:
                    line += f" (Comment: {cell.comment.text.strip()})"

                row_lines.append(line)

            if row_has_data:
                output_lines.extend(row_lines)

    return "\n".join(output_lines)


if __name__ == "__main__":
    file_path = "highlighted_example.xlsx"
    
    # Historical → ignore dropdown options, but include colors
    text_output = excel_to_table_text(file_path, historical=True)
    print(text_output)

    with open("processed_complex_format_empty.txt", "w", encoding="utf-8") as f:
        f.write(text_output)

    print("Processed Excel → Table-like text empty saved as processed_complex_format_empty.txt")