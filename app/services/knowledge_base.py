from typing import TypedDict, List, Dict, Any, Annotated, Optional, Tuple
from langgraph.graph import StateGraph, END
from openpyxl import load_workbook
from app.services.excel_parser import sheet_to_table_text
from concurrent.futures import ThreadPoolExecutor
from app.core.clients.bedrock import bedrock_client
from app.core.clients.qdrant import ensure_collection, create_embeddings, qdrant, COLLECTION_NAME
from qdrant_client.models import PointStruct
from app.core.logger import get_logger
from pathlib import Path
import json
import re
import uuid

logger = get_logger(__name__)


def merge_dicts(x: Dict[str, Any], y: Dict[str, Any]) -> Dict[str, Any]:
    """Merge two dictionaries."""
    return {**x, **y}


class SheetProcessingState(TypedDict):
    """State for individual sheet processing subgraph."""
    workbook_path: str
    sheet_name: str
    table_text: str
    batches: List[Dict[str, Any]]  # List of batches, each with row_numbers and structure_info
    batch_facts: Dict[int, List[str]]  # Dictionary mapping batch index to list of fact paragraphs
    sectors: List[str]  # Sectors for context
    technologies: List[str]  # Technologies for context
    error: str


class KnowledgeBaseState(TypedDict):
    """State for the knowledge base processing flow."""
    sectors: List[str]
    technologies: List[str]
    workbook_path: str
    sheet_names: List[str]
    sheet_results: Annotated[Dict[str, Any], merge_dicts]  # Each value is a dict with "table_text", "batches", and "batch_facts"
    errors: List[str]
    facts_upserted: bool  # Flag to track if facts have been upserted to Qdrant


def extract_sectors_and_technologies_from_filename(filepath: str) -> Tuple[List[str], List[str]]:
    """
    Extract sector and technology names from workbook filename.
    
    Format: [client name]-[sector]_[technology1]_[technology2]
    - Client name is separated by "-"
    - First part after "-" is the sector
    - Remaining parts separated by "_" are technologies
    - Removes "sector"/"sect" from sector names
    - Removes "technology"/"tech" from technology names
    
    Args:
        filepath: Path to the workbook file
        
    Returns:
        Tuple of (sectors list, technologies list)
        
    Raises:
        ValueError: If the filename format is invalid or sectors/technologies cannot be parsed
    """
    # Get filename without extension
    filename = Path(filepath).stem
    
    # Remove curly braces and their contents from filename
    # Example: "ClientName-{something}Sector_Technology" -> "ClientName-Sector_Technology"
    filename = re.sub(r'\{[^}]*\}', '', filename)
    
    # Split by "-" to separate client name from sector/technologies
    if "-" not in filename:
        raise ValueError(
            f"Invalid filename format. Expected format: [client name]-[sector]_[technology1]_[technology2]. "
            f"Filename: {filename}. Missing '-' separator."
        )
    
    parts = filename.split("-", 1)  # Split only on first "-"
    if len(parts) < 2:
        raise ValueError(
            f"Invalid filename format. Expected format: [client name]-[sector]_[technology1]_[technology2]. "
            f"Filename: {filename}. Could not split by '-'."
        )
    
    # parts[0] is client name, parts[1] contains sector and technologies
    sector_and_techs = parts[1]
    
    # Split by "_" to get sector (first) and technologies (rest)
    components = sector_and_techs.split("_")
    
    if not components:
        raise ValueError(
            f"Invalid filename format. Expected format: [client name]-[sector]_[technology1]_[technology2]. "
            f"Filename: {filename}. No components found after '-'."
        )
    
    # First component is the sector
    sector = components[0].strip()
    
    if not sector:
        raise ValueError(
            f"Invalid filename format. Expected format: [client name]-[sector]_[technology1]_[technology2]. "
            f"Filename: {filename}. Sector name is empty."
        )
    
    # Remove "sector" or "sect" from sector name (case-insensitive)
    sector = re.sub(r'\b(sector|sect)\b', '', sector, flags=re.IGNORECASE).strip()
    # Clean up any extra spaces or underscores
    sector = re.sub(r'[_\s]+', ' ', sector).strip()
    
    if not sector:
        raise ValueError(
            f"Invalid filename format. Expected format: [client name]-[sector]_[technology1]_[technology2]. "
            f"Filename: {filename}. Sector name is empty after cleaning."
        )
    
    sectors = [sector]
    
    # Remaining components are technologies
    technologies = []
    for tech in components[1:]:
        tech = tech.strip()
        if tech:
            # Remove "technology" or "tech" from technology name (case-insensitive)
            tech = re.sub(r'\b(technology|tech)\b', '', tech, flags=re.IGNORECASE).strip()
            # Clean up any extra spaces or underscores
            tech = re.sub(r'[_\s]+', ' ', tech).strip()
            if tech:
                technologies.append(tech)
    
    if not technologies:
        raise ValueError(
            f"Invalid filename format. Expected format: [client name]-[sector]_[technology1]_[technology2]. "
            f"Filename: {filename}. No technologies found after sector."
        )
    
    return sectors, technologies


def initialize_workbook(state: KnowledgeBaseState) -> KnowledgeBaseState:
    """
    Initialize the workbook and extract sheet names, sectors, and technologies.
    
    Args:
        state: Current state
        
    Returns:
        Updated state with sheet names, sectors, and technologies
        
    Raises:
        ValueError: If sectors and technologies cannot be parsed from filename
    """
    try:
        # Extract sectors and technologies from filename - this will raise ValueError if parsing fails
        sectors, technologies = extract_sectors_and_technologies_from_filename(state["workbook_path"])
        state["sectors"] = sectors
        state["technologies"] = technologies
        
        # Log extracted sectors and technologies
        logger.info(f"Extracted sectors from filename: {sectors}")
        logger.info(f"Extracted technologies from filename: {technologies}")
        print(f"📋 Extracted Sectors: {sectors}")
        print(f"🔧 Extracted Technologies: {technologies}")
        
        # Load workbook and extract sheet names (only visible sheets)
        workbook = load_workbook(state["workbook_path"], data_only=True)
        sheet_names = [
            name for name in workbook.sheetnames
            if workbook[name].sheet_state == "visible"
        ]
        
        state["sheet_names"] = sheet_names
        state["sheet_results"] = {}
        state["errors"] = []
        state["facts_upserted"] = False
        
        logger.info(f"Initialized workbook with {len(sheet_names)} visible sheets")
        print(f"📊 Found {len(sheet_names)} visible sheets: {sheet_names}")
        
    except ValueError as e:
        # Log the error before re-raising
        error_msg = str(e)
        logger.error(f"Failed to extract sectors/technologies from filename: {error_msg}")
        print(f"❌ Error extracting sectors/technologies: {error_msg}")
        state["errors"].append(error_msg)
        raise
    except Exception as e:
        # Log any other errors
        error_msg = f"Error initializing workbook: {str(e)}"
        logger.error(error_msg, exc_info=True)
        print(f"❌ {error_msg}")
        state["errors"].append(error_msg)
        raise
    
    return state


def process_single_sheet(state: SheetProcessingState) -> SheetProcessingState:
    """
    Process a single sheet and extract table text.
    This is used within the sheet processing subgraph.
    
    Args:
        state: Sheet processing state
        
    Returns:
        Updated state with table text
    """
    sheet_name = state.get("sheet_name", "unknown")
    logger.info(f"Starting to process sheet: {sheet_name}")
    print(f"📄 Processing sheet: {sheet_name}")
    
    try:
        logger.debug(f"Loading workbook: {state['workbook_path']}")
        workbook = load_workbook(state["workbook_path"], data_only=True)
        logger.debug(f"Workbook loaded, accessing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        logger.debug(f"Extracting table text from sheet: {sheet_name}")
        table_text = sheet_to_table_text(sheet, sheet_name, historical=True)
        logger.info(f"Extracted {len(table_text)} characters of table text from sheet: {sheet_name}")
        print(f"  ✓ Extracted table text: {len(table_text)} characters")
        state["table_text"] = table_text
        state["batches"] = []
        state["batch_facts"] = {}
        # sectors and technologies should already be in state
        state["error"] = ""
    except Exception as e:
        error_msg = f"Error processing sheet {sheet_name}: {str(e)}"
        logger.error(error_msg, exc_info=True)
        print(f"  ❌ {error_msg}")
        state["table_text"] = ""
        state["batches"] = []
        state["batch_facts"] = {}
        state["error"] = error_msg
    
    return state


def analyze_table_batches(state: SheetProcessingState) -> SheetProcessingState:
    """
    Analyze table text using LLM to create batches of questions.
    Each batch contains row numbers for a single independent question,
    its answer rows, and follow-up questions/answer rows.
    
    Args:
        state: Sheet processing state with table_text
        
    Returns:
        Updated state with batches
    """
    sheet_name = state.get("sheet_name", "unknown")
    
    if not state.get("table_text") or state.get("error"):
        logger.warning(f"Skipping batch analysis for sheet {sheet_name}: no table_text or error present")
        return state
    
    logger.info(f"Starting batch analysis for sheet: {sheet_name}")
    print(f"  🔍 Analyzing batches for sheet: {sheet_name}")
    
    try:
        system_message = """You are an expert Excel analyzer that identifies question batches in risk questionnaires.

Your task is to analyze the provided Excel table text and identify batches of questions. Each batch represents:
1. A single independent question with its answer rows
2. All follow-up questions that depend on that question's answers, along with their answer rows

CRITICAL FILTERING RULES - MUST BE STRICTLY FOLLOWED:
- ONLY create batches for questions that HAVE ANSWERS PROVIDED and that is not just a reference to an attachment/document with NO other content
- DO NOT create batches for questions with NO answer (empty/null/blank cells)
- DO NOT create batches for questions where the answer ONLY references an attachment/document with NO other substantive content
- Before creating a batch, verify that the question has at least one answer cell with a value
- If a question row exists but all answer cells are empty or marked as "(empty)", DO NOT create a batch for it
- IMPORTANT: Include questions with ANY answer value, including "N/A", "Not Applicable", "None", etc. - these are valid answers
- CRITICAL: If an answer cell contains ONLY phrases like "see attached PDF", "refer to document", "see attachment", "please see PDF", "see attached", "refer to PDF", "see document", "please see document", "see the attached file", "refer to attached document" with NO other content, DO NOT create a batch for that question
- ONLY include questions about COMPLIANCE and RISK for the organization
- EXCLUDE questions about personal information such as:
  * Who filled the sheet / Name of person
  * When the sheet was filled / Date / Timestamp
  * Contact information / Email / Phone
  * Job title / Department (unless directly related to risk/compliance)
  * Any other personal or administrative metadata
- EXCLUDE questions that are purely informational or administrative (not about compliance/risk)

IMPORTANT RULES:
- Each batch must contain ALL rows related to one independent question (question row + answer rows + follow-up question rows + follow-up answer rows)
- Follow-up questions are questions that depend on the main question's answers
- Only include follow-up questions if they also have answers (not empty)
- All answer rows and question rows for a complete question set should be in the same batch
- Return ONLY valid JSON. Do not include any text before or after the JSON.

OUTPUT FORMAT (return ONLY JSON):
{
  "batches": [
    {
      "row_numbers": [5, 6, 7, 8, 9],
      "structure_info": "Question row: 5 (cells A5-D5), Answer rows: 6-7 (cells A6-D7), Follow-up question: 8 (cells A8-D8), Follow-up answer: 9 (cells A9-D9)"
    },
    {
      "row_numbers": [12, 13, 14],
      "structure_info": "Question row: 12 (cells B12-E12), Answer rows: 13-14 (cells B13-E14)"
    }
  ]
}

For each batch, provide:
- row_numbers: List of all row numbers (integers) that belong to this batch
- structure_info: A one-line description of the structure in terms of cell numbers and row numbers for the complete batch"""

        prompt = f"""Analyze the following Excel table text and identify batches of questions.

SHEET: {state['sheet_name']}

TABLE TEXT:
{state['table_text']}

CRITICAL: Apply strict filtering before creating batches:
1. ONLY create batches for questions that HAVE ANSWERS PROVIDED
   - DO NOT create batches for questions with NO answer (empty/null/blank cells)
   - Before including a question in a batch, verify it has at least one answer cell with a value
   - Look for cells with patterns like: CELL = "value" (this has an answer)
   - Skip cells with patterns like: CELL = (empty) (this has NO answer)
   - IMPORTANT: Include questions with ANY answer value, even if it's "N/A", "Not Applicable", "None", etc. - these are considered valid answers
   - If you find a question row but all its answer cells are empty or marked as "(empty)", DO NOT create a batch for it
2. DO NOT create batches for questions where the answer ONLY references an attachment/document
   - If an answer cell contains ONLY phrases like "see attached PDF", "refer to document", "see attachment", "please see PDF", "see attached", "refer to PDF", "see document", "please see document", "see the attached file", "refer to attached document" with NO other substantive content, DO NOT create a batch for that question
   - Check each answer cell carefully - if it ONLY contains attachment references and nothing else, skip that question entirely
   - DO NOT include such questions in any batch
3. ONLY include questions about COMPLIANCE and RISK for the organization
4. EXCLUDE questions about:
   - Who filled the sheet / Name / Person information
   - When it was filled / Date / Timestamp
   - Contact information / Email / Phone
   - Job title / Department (unless directly related to risk/compliance assessment)
   - Any personal or administrative metadata

For questions that pass ALL filters AND have substantive answers (not just attachment references), group them into batches. Each batch should contain:
- The main question row (only if it has answer rows with substantive values, not just attachment references)
- All answer rows for that question (including rows with "N/A" or similar values, but NOT empty rows and NOT attachment-only references)
- All follow-up question rows (if any) - but only if they also have substantive answers (including "N/A", NOT empty, NOT attachment-only)
- All follow-up answer rows (if any) - only if they contain actual substantive values (not just attachment references)

For each batch, provide:
1. row_numbers: A list of all row numbers (as integers) that belong to this batch
2. structure_info: A one-line description explaining the structure in terms of cell numbers and row numbers

Return ONLY valid JSON matching this format:
{{
  "batches": [
    {{
      "row_numbers": [5, 6, 7, 8, 9],
      "structure_info": "Question row: 5 (cells A5-D5), Answer rows: 6-7 (cells A6-D7), Follow-up question: 8 (cells A8-D8), Follow-up answer: 9 (cells A9-D9)"
    }}
  ]
}}"""

        # Call LLM to analyze and create batches
        logger.debug(f"Calling LLM for batch analysis on sheet: {sheet_name}")
        print(f"    → Calling LLM for batch analysis...")
        raw_output = bedrock_client.invoke_with_tracing(
            prompt=prompt,
            system_message=system_message
        )
        logger.debug(f"LLM response received for sheet: {sheet_name}, length: {len(raw_output)}")
        print(f"    ✓ LLM response received ({len(raw_output)} chars)")
        
        # Clean and parse the output
        logger.debug(f"Cleaning LLM output for sheet: {sheet_name}")
        cleaned_output = _clean_llm_output(raw_output)
        
        # Try to parse JSON
        logger.debug(f"Parsing JSON output for sheet: {sheet_name}")
        try:
            parsed_output = json.loads(cleaned_output)
            batches = parsed_output.get("batches", [])
            logger.info(f"Parsed {len(batches)} batches from sheet: {sheet_name}")
            print(f"    ✓ Parsed {len(batches)} batches")
            
            # Validate batches structure
            validated_batches = []
            for batch in batches:
                if "row_numbers" in batch and "structure_info" in batch:
                    # Ensure row_numbers is a list of integers
                    row_nums = batch["row_numbers"]
                    if isinstance(row_nums, list):
                        validated_batches.append({
                            "row_numbers": [int(r) for r in row_nums if isinstance(r, (int, str)) and str(r).isdigit()],
                            "structure_info": str(batch["structure_info"])
                        })
            
            state["batches"] = validated_batches
            state["error"] = ""
            logger.info(f"Validated {len(validated_batches)} batches for sheet: {sheet_name}")
            print(f"  ✓ Created {len(validated_batches)} validated batches")
            
        except json.JSONDecodeError as e:
            error_msg = f"Error parsing LLM response as JSON for sheet {sheet_name}: {str(e)}. Raw output: {cleaned_output[:500]}"
            logger.warning(error_msg)
            print(f"  ⚠️  {error_msg}")
            state["error"] = error_msg
            state["batches"] = []
        except Exception as e:
            error_msg = f"Error processing batches for sheet {sheet_name}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            print(f"  ❌ {error_msg}")
            state["error"] = error_msg
            state["batches"] = []
            
    except Exception as e:
        error_msg = f"Error analyzing table batches for sheet {sheet_name}: {str(e)}"
        logger.error(error_msg, exc_info=True)
        print(f"  ❌ {error_msg}")
        state["error"] = error_msg
        state["batches"] = []
    
    return state


def _clean_llm_output(raw_output: str) -> str:
    """Clean and format the LLM output."""
    # Remove code fences if present
    cleaned = re.sub(r"^```[a-zA-Z]*\n?", "", raw_output.strip(), flags=re.MULTILINE)
    cleaned = re.sub(r"\n?```$", "", cleaned).strip()
    
    # Remove reasoning blocks if present
    reasoning_pattern = r'<reasoning>.*?</reasoning>'
    cleaned = re.sub(reasoning_pattern, '', cleaned, flags=re.DOTALL)
    
    # Try to extract JSON if wrapped in text
    json_match = re.search(r'\{.*\}', cleaned, re.DOTALL)
    if json_match:
        cleaned = json_match.group(0)
    
    # Clean up any extra whitespace
    cleaned = re.sub(r'\n\s*\n\s*\n', '\n\n', cleaned)
    
    return cleaned.strip()


def extract_row_table_text(table_text: str, row_number: int) -> str:
    """
    Extract table text from a specific row to just before the next row marker.
    Extracts from "Row x" to just before the next "Row" marker (not necessarily "Row x+1").
    
    Args:
        table_text: Full table text from the sheet
        row_number: The row number to extract (e.g., 25)
        
    Returns:
        Extracted table text from "Row x" to just before the next "Row" marker
    """
    if not table_text or row_number < 1:
        return ""
    
    # Split table text into lines
    lines = table_text.split('\n')
    
    # Find the start line (Row x)
    start_idx = None
    for i, line in enumerate(lines):
        if line.strip().startswith(f"Row {row_number}"):
            start_idx = i
            break
    
    if start_idx is None:
        # If exact match not found, return empty string
        return ""
    
    # Find the end line (next Row marker, not necessarily Row x+1)
    end_idx = len(lines)
    
    # Look for the next row marker (any row number greater than current)
    for i in range(start_idx + 1, len(lines)):
        line = lines[i].strip()
        # Check if this line starts with "Row " followed by a number
        if line.startswith("Row "):
            end_idx = i
            break
    
    # Extract the relevant lines
    extracted_lines = lines[start_idx:end_idx]
    return '\n'.join(extracted_lines)


def extract_batch_table_text(table_text: str, row_numbers: List[int]) -> str:
    """
    Extract table text for all rows in a batch by extracting each row
    and joining them together.
    
    Args:
        table_text: Full table text from the sheet
        row_numbers: List of row numbers in the batch (e.g., [25, 26, 27])
        
    Returns:
        Combined table text for all rows in the batch
    """
    if not row_numbers:
        return ""
    
    # Extract text for each row and join them
    extracted_texts = []
    for row_num in sorted(row_numbers):
        row_text = extract_row_table_text(table_text, row_num)
        if row_text:
            extracted_texts.append(row_text)
    
    return '\n\n'.join(extracted_texts)


def convert_batch_to_facts(batch_text: str, batch_idx: int, sheet_name: str, sectors: List[str], technologies: List[str]) -> List[str]:
    """
    Convert batch table text into fact-like paragraphs using LLM.
    Creates one paragraph per question (main question + follow-ups).
    
    Args:
        batch_text: Combined table text for the batch
        batch_idx: Index of the batch
        sheet_name: Name of the sheet
        sectors: List of sectors
        technologies: List of technologies
        
    Returns:
        List of fact paragraphs (one per question)
    """
    if not batch_text:
        return []
    
    try:
        sectors_str = ", ".join(sectors) if sectors else "Not specified"
        technologies_str = ", ".join(technologies) if technologies else "Not specified"
        
        system_message = """You are an expert at converting risk questionnaire data into structured fact paragraphs for vector database storage.

Your task is to convert question-answer pairs from risk questionnaires into single, self-sufficient fact paragraphs optimized for vector database retrieval.

CRITICAL RULE - MUST BE FOLLOWED STRICTLY:
- ONE question with a valid answer = EXACTLY ONE fact paragraph
- NO EXCEPTIONS: Even if the answer spans multiple rows, multiple cells, or is very long, it MUST be consolidated into ONE single paragraph
- DO NOT create separate paragraphs for different rows of the same answer
- DO NOT create separate paragraphs for different parts of the same answer
- The entire answer, regardless of length or structure, should be summarized/described in ONE paragraph

KEY PRINCIPLES:
1. One question-answer pair = One fact paragraph (ALWAYS)
2. Each paragraph must be self-sufficient and complete (include all context needed for retrieval)
3. Focus on Compass Group's operations, technologies, processes, compliance, and risk management
4. Any vendor mention refers to "Compass Group" (the organization answering the questionnaire)
5. Exclude personal information, sheet metadata, and attachment-only references

OUTPUT FORMAT:
Return facts as plain text paragraphs, one per line, without numbering or formatting."""

        prompt = f"""Convert the following risk questionnaire table text into fact paragraphs.

SHEET: {sheet_name}
BATCH INDEX: {batch_idx}

CONTEXT:
- SECTORS: {sectors_str}
  (Sectors refer to departments within Compass Group)
  NOTE: One technology may be used in more than one sector. Do NOT mention sectors in facts.
- TECHNOLOGIES/APPLICATIONS: {technologies_str}
  (Technologies refer to products/services that Compass Group provides to their clients)
  NOTE: Facts should focus on technologies/applications, not sectors.

TABLE TEXT:
{batch_text}

STEP 1: Identify all questions in this batch (main question + any follow-up questions) and their count

STEP 2: For EACH question, create EXACTLY ONE fact paragraph. CRITICAL: Even if the answer spans multiple rows, multiple cells, or is very long, you MUST consolidate everything into ONE single paragraph:

A. N/A ANSWERS (MUST CREATE FACT):
   - If answer is "N/A", "Not Applicable", "NA", "n/a", "not applicable", or any variation
   - Create ONE fact paragraph stating what is not applicable, including question context
   - Example: Question "Do you use entity-developed spreadsheets?" with answer "N/A" 
     → Fact: "Entity-developed spreadsheets are not applicable for Compass Group."
   - DO NOT skip N/A answers - they are valid answers

B. ATTACHMENT-ONLY ANSWERS (SKIP - NO FACT):
   - If answer ONLY mentions some reference to an attachment/document with NO other content
   - Examples: "see attached PDF", "refer to document", "see attachment", "please see PDF", "see attached", "refer to PDF", "see document", "please see document", "see the attached file", "refer to attached document"
   - DO NOT create a fact - skip this question entirely and just return empty string for that question as fact paragraph

C. ANSWERS WITH ATTACHMENT + SUBSTANTIVE CONTENT:
   - Extract ONLY the substantive part
   - Completely ignore and remove the attachment reference
   - Create ONE fact paragraph with only the substantive content

D. MULTI-ROW ANSWERS (ONE FACT FOR ENTIRE ANSWER):
   - If answer spans multiple rows (table structure, list, or any multi-row format)
   - This is ONE answer to ONE question, regardless of how many rows it spans
   - Example: Question row 9, Answer rows 10-15 = ONE question = ONE paragraph
   - Create ONE paragraph that describes/summarizes the ENTIRE answer (all rows together)
   - DO NOT create separate paragraphs for each row - all rows together = one answer = one paragraph
   - Consolidate all information from all rows into a single coherent paragraph

E. REGULAR ANSWERS (SINGLE OR MULTI-ROW):
   - Create ONE fact paragraph with the answer content
   - If answer spans multiple rows, consolidate all rows into ONE paragraph
   - Include question context to make the fact self-sufficient
   - DO NOT split into multiple paragraphs based on row count

STEP 3: Output rules - STRICTLY ENFORCE ONE PARAGRAPH PER QUESTION:
- ONE question with valid answer = EXACTLY ONE paragraph (NO EXCEPTIONS)
- Main question = 1 paragraph, Each follow-up question = 1 paragraph
- Count questions, NOT rows: 1 question with 10 rows of answer = 1 paragraph
- If answer has multiple rows, consolidate ALL rows into ONE paragraph
- DO NOT create multiple paragraphs for the same question, regardless of answer length or structure
- Return ONLY plain text paragraphs, one per line, no numbering or formatting
- Exclude personal info, sheet metadata, and attachment-only references
- Focus on Compass Group's operations, technologies, processes, compliance, and risk management
- IMPORTANT: Do NOT mention sectors in facts - focus only on technologies/applications
- Since one technology may be used across multiple sectors, facts should be technology-focused, not sector-specific

CRITICAL REMINDER: 
- ONE question = ONE paragraph, ALWAYS
- Multiple answer rows = ONE consolidated paragraph
- Long answers = ONE comprehensive paragraph
- Table answers = ONE summarizing paragraph
- NO splitting, NO multiple paragraphs per question"""

        # Call LLM to convert to facts
        logger.debug(f"Calling LLM to convert batch {batch_idx} to facts for sheet: {sheet_name}")
        raw_output = bedrock_client.invoke_with_tracing(
            prompt=prompt,
            system_message=system_message
        )
        logger.debug(f"LLM response received for batch {batch_idx}, length: {len(raw_output)}")
        
        # Clean the output
        cleaned_output = _clean_llm_output(raw_output)
        
        # Split into paragraphs (split by newlines and filter empty)
        # Note: The LLM should return one paragraph per question, but we split by newlines
        # in case the LLM returns multiple lines for formatting
        paragraphs = [p.strip() for p in cleaned_output.split('\n') if p.strip()]
        
        # Filter out formatting artifacts but keep all substantive paragraphs
        # The LLM should have already created one paragraph per question
        facts = [p for p in paragraphs if len(p) > 10 and not p.startswith(('1.', '2.', '3.', '4.', '5.', '-', '*', '•', 'A.', 'B.', 'C.', 'D.', 'E.'))]
        
        # If we have multiple paragraphs but they might be from the same question,
        # we trust the LLM's output. However, if there are too many paragraphs relative
        # to expected questions, we log a warning.
        logger.debug(f"Generated {len(facts)} facts for batch {batch_idx} in sheet: {sheet_name}")
        
        if len(facts) > 10:  # Arbitrary threshold - if we have many paragraphs, might indicate splitting issue
            logger.warning(f"Batch {batch_idx} generated {len(facts)} facts - might indicate multiple paragraphs per question")
        
        return facts
        
    except Exception as e:
        error_msg = f"Error converting batch {batch_idx} to facts for sheet {sheet_name}: {str(e)}"
        logger.error(error_msg, exc_info=True)
        print(f"    ❌ {error_msg}")
        return []


def process_batches_to_facts(state: SheetProcessingState) -> SheetProcessingState:
    """
    Process batches and convert them into fact paragraphs.
    Processes batches sequentially to avoid rate limits.
    
    Args:
        state: Sheet processing state with batches, table_text, sectors, and technologies
        
    Returns:
        Updated state with batch_facts populated
    """
    sheet_name = state.get("sheet_name", "unknown")
    
    if not state.get("batches") or not state.get("table_text") or state.get("error"):
        logger.warning(f"Skipping fact conversion for sheet {sheet_name}: no batches or error present")
        state["batch_facts"] = {}
        return state
    
    try:
        batch_facts = {}
        table_text = state["table_text"]
        batches = state["batches"]
        sectors = state.get("sectors", [])
        technologies = state.get("technologies", [])
        
        logger.info(f"Starting fact conversion for {len(batches)} batches in sheet: {sheet_name}")
        print(f"  📝 Converting {len(batches)} batches to facts for sheet: {sheet_name}")
        
        # Process batches sequentially to avoid rate limits
        for batch_idx, batch in enumerate(batches):
            logger.debug(f"Processing batch {batch_idx + 1}/{len(batches)} for sheet: {sheet_name}")
            print(f"    → Processing batch {batch_idx + 1}/{len(batches)}...")
            row_numbers = batch.get("row_numbers", [])
            if not row_numbers:
                batch_facts[batch_idx] = []
                continue
            
            # Extract table text for all rows in the batch
            logger.debug(f"Extracting table text for batch {batch_idx + 1} (rows: {row_numbers})")
            batch_text = extract_batch_table_text(table_text, row_numbers)
            logger.debug(f"Extracted {len(batch_text)} characters for batch {batch_idx + 1}")
            
            # Convert to facts (one paragraph per question)
            logger.debug(f"Converting batch {batch_idx + 1} to facts")
            facts = convert_batch_to_facts(batch_text, batch_idx, sheet_name, sectors, technologies)
            batch_facts[batch_idx] = facts
            logger.info(f"Batch {batch_idx + 1}/{len(batches)} converted: {len(facts)} facts")
            print(f"    ✓ Batch {batch_idx + 1}/{len(batches)}: {len(facts)} facts")
        
        logger.info(f"Completed fact conversion for sheet {sheet_name}: {len(batch_facts)} batches processed")
        print(f"  ✓ Completed fact conversion: {sum(len(f) for f in batch_facts.values())} total facts")
        state["batch_facts"] = batch_facts
        
        if not state.get("error"):
            state["error"] = ""
            
    except Exception as e:
        state["error"] = f"Error converting batches to facts: {str(e)}"
        state["batch_facts"] = {}
    
    return state


def save_facts_to_file(state: SheetProcessingState, output_dir: Path) -> None:
    """
    Save batch facts to a text file.
    
    Args:
        state: Sheet processing state with batch_facts
        output_dir: Directory to save the file
    """
    if not state.get("batch_facts"):
        return
    
    try:
        sheet_name = state.get("sheet_name", "unknown")
        # Create a safe filename from sheet name
        safe_sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_sheet_name = safe_sheet_name.replace(' ', '_')
        
        output_file = output_dir / f"facts_{safe_sheet_name}.txt"
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"Facts for Sheet: {sheet_name}\n")
            f.write("=" * 80 + "\n\n")
            
            for batch_idx in sorted(state["batch_facts"].keys()):
                facts = state["batch_facts"][batch_idx]
                if facts:
                    f.write(f"Batch {batch_idx + 1}:\n")
                    f.write("-" * 80 + "\n")
                    for fact in facts:
                        f.write(f"{fact}\n\n")
                    f.write("\n")
        
        print(f"💾 Facts saved to: {output_file}")
        
    except Exception as e:
        print(f"⚠️  Warning: Could not save facts to file: {e}")


def create_sheet_subgraph() -> StateGraph:
    """
    Create a subgraph workflow for processing a single sheet.
    
    Returns:
        Compiled subgraph for sheet processing
    """
    subgraph = StateGraph[SheetProcessingState, None, SheetProcessingState, SheetProcessingState](SheetProcessingState)
    
    # Add the processing nodes
    subgraph.add_node("process_sheet", process_single_sheet)
    subgraph.add_node("analyze_batches", analyze_table_batches)
    subgraph.add_node("convert_to_facts", process_batches_to_facts)
    
    # Define the flow: extract table text, analyze batches, then convert to facts
    subgraph.set_entry_point("process_sheet")
    subgraph.add_edge("process_sheet", "analyze_batches")
    subgraph.add_edge("analyze_batches", "convert_to_facts")
    subgraph.add_edge("convert_to_facts", END)
    
    return subgraph.compile()


def _invoke_sheet_subgraph(sheet_subgraph, workbook_path: str, sheet_name: str, sectors: List[str], technologies: List[str]) -> tuple[str, SheetProcessingState]:
    """
    Invoke the sheet processing subgraph for a single sheet.
    
    Args:
        sheet_subgraph: The compiled sheet processing subgraph
        workbook_path: Path to the workbook
        sheet_name: Name of the sheet to process
        sectors: List of sectors
        technologies: List of technologies
        
    Returns:
        Tuple of (sheet_name, result_state)
    """
    logger.info(f"Invoking sheet subgraph for: {sheet_name}")
    print(f"  🔄 Starting subgraph for sheet: {sheet_name}")
    
    sheet_state: SheetProcessingState = {
        "workbook_path": workbook_path,
        "sheet_name": sheet_name,
        "table_text": "",
        "batches": [],
        "batch_facts": {},
        "sectors": sectors,
        "technologies": technologies,
        "error": ""
    }
    logger.debug(f"Invoking subgraph with state for sheet: {sheet_name}")
    result = sheet_subgraph.invoke(sheet_state)
    logger.info(f"Subgraph completed for sheet: {sheet_name}")
    print(f"  ✓ Subgraph completed for sheet: {sheet_name}")
    return (sheet_name, result)


def process_all_sheets(state: KnowledgeBaseState) -> KnowledgeBaseState:
    """
    Process all sheets using subgraphs for each sheet in parallel.
    
    Args:
        state: Current state
        
    Returns:
        Updated state with sheet results
    """
    if not state.get("sheet_names"):
        logger.warning("No sheet names found, skipping sheet processing")
        return state
    
    workbook_path = state["workbook_path"]
    sheet_names = state["sheet_names"]
    sheet_results = {}
    errors = state.get("errors", [])
    
    logger.info(f"Starting to process {len(sheet_names)} sheets: {sheet_names}")
    print(f"\n🚀 Starting to process {len(sheet_names)} sheets...")
    
    # Create the sheet processing subgraph
    logger.debug("Creating sheet processing subgraph")
    sheet_subgraph = create_sheet_subgraph()
    logger.debug("Sheet processing subgraph created")
    
    # Process all sheets in parallel using ThreadPoolExecutor
    sectors = state.get("sectors", [])
    technologies = state.get("technologies", [])
    
    logger.info(f"Submitting {len(sheet_names)} sheet processing tasks to ThreadPoolExecutor")
    print(f"  📋 Submitting {len(sheet_names)} sheets for processing...")
    
    with ThreadPoolExecutor(max_workers=len(sheet_names)) as executor:
        # Submit all sheet processing tasks
        futures = [
            executor.submit(_invoke_sheet_subgraph, sheet_subgraph, workbook_path, sheet_name, sectors, technologies)
            for sheet_name in state["sheet_names"]
        ]
        
        # Collect results as they complete and save facts to file
        output_dir = Path(state["workbook_path"]).parent / "facts_output"
        output_dir.mkdir(exist_ok=True)
        logger.info(f"Output directory: {output_dir}")
        
        logger.info(f"Waiting for {len(futures)} sheet processing tasks to complete...")
        print(f"  ⏳ Waiting for sheet processing to complete...")
        
        for idx, future in enumerate(futures, 1):
            logger.debug(f"Waiting for sheet {idx}/{len(futures)} to complete...")
            print(f"    → Waiting for sheet {idx}/{len(futures)}...")
            sheet_name, result = future.result()
            logger.info(f"Sheet {idx}/{len(futures)} completed: {sheet_name}")
            print(f"    ✓ Sheet {idx}/{len(futures)} completed: {sheet_name}")
            
            if result["error"]:
                errors.append(result["error"])
                logger.error(f"Error in sheet {sheet_name}: {result['error']}")
            else:
                # Save facts to file
                logger.debug(f"Saving facts to file for sheet: {sheet_name}")
                save_facts_to_file(result, output_dir)
                logger.debug(f"Facts saved for sheet: {sheet_name}")
                
                # Store table_text, batches, and batch_facts for each sheet
                sheet_results[sheet_name] = {
                    "table_text": result["table_text"],
                    "batches": result.get("batches", []),
                    "batch_facts": result.get("batch_facts", {})
                }
    
    # Update main state
    state["sheet_results"] = sheet_results
    state["errors"] = errors
    
    # Log errors if any
    if errors:
        logger.warning(f"Encountered {len(errors)} error(s) during sheet processing")
        for error in errors:
            logger.error(f"Sheet processing error: {error}")
            print(f"⚠️  Error: {error}")
    else:
        logger.info("All sheets processed successfully with no errors")
    
    return state


def upsert_facts_to_qdrant(state: KnowledgeBaseState) -> KnowledgeBaseState:
    """
    Collect all facts from all sheets and upsert them into Qdrant with metadata.
    Each fact paragraph is treated as one chunk with sectors and technologies as metadata.
    
    Args:
        state: Current state with sheet_results containing batch_facts
        
    Returns:
        Updated state with facts_upserted flag set to True
    """
    logger.info("Starting to upsert facts to Qdrant")
    print(f"\n💾 Upserting facts to Qdrant...")
    
    try:
        # Ensure collection exists
        ensure_collection()
        logger.debug("Qdrant collection ensured")
        
        # Collect all facts from all sheets
        all_facts = []
        sectors = state.get("sectors", [])
        technologies = state.get("technologies", [])
        
        sheet_results = state.get("sheet_results", {})
        logger.info(f"Collecting facts from {len(sheet_results)} sheets")
        
        for sheet_name, sheet_data in sheet_results.items():
            if isinstance(sheet_data, dict):
                batch_facts = sheet_data.get("batch_facts", {})
                if batch_facts:
                    for batch_idx, facts in batch_facts.items():
                        for fact in facts:
                            if fact and fact.strip():  # Only add non-empty facts
                                all_facts.append(fact.strip())
        
        if not all_facts:
            logger.warning("No facts found to upsert to Qdrant")
            print(f"  ⚠️  No facts found to upsert")
            state["facts_upserted"] = False
            return state
        
        logger.info(f"Collected {len(all_facts)} facts to upsert")
        print(f"  📝 Collected {len(all_facts)} facts from all sheets")
        
        # Create embeddings for all facts
        logger.info("Creating embeddings for facts")
        print(f"  🔄 Creating embeddings...")
        embeddings = create_embeddings(all_facts)
        logger.info(f"Created {len(embeddings)} embeddings")
        print(f"  ✓ Created {len(embeddings)} embeddings")
        
        # Prepare points with metadata
        points = []
        for fact, embedding in zip(all_facts, embeddings):
            points.append(
                PointStruct(
                    id=str(uuid.uuid4()),
                    vector=embedding,
                    payload={
                        "text": fact,
                        "sectors": sectors,  # List of sectors for filtering
                        "technologies": technologies  # List of technologies for filtering
                    }
                )
            )
        
        # Upsert to Qdrant
        logger.info(f"Upserting {len(points)} points to Qdrant")
        print(f"  🔄 Upserting {len(points)} chunks to Qdrant...")
        qdrant.upsert(
            collection_name=COLLECTION_NAME,
            points=points
        )
        
        logger.info(f"Successfully upserted {len(points)} facts to Qdrant collection '{COLLECTION_NAME}'")
        print(f"  ✅ Successfully upserted {len(points)} facts to Qdrant")
        print(f"     - Metadata: sectors={sectors}, technologies={technologies}")
        
        state["facts_upserted"] = True
        
    except Exception as e:
        error_msg = f"Error upserting facts to Qdrant: {str(e)}"
        logger.error(error_msg, exc_info=True)
        print(f"  ❌ {error_msg}")
        state["facts_upserted"] = False
        state["errors"].append(error_msg)
    
    return state


# Create the LangGraph workflow
def create_knowledge_base_workflow() -> StateGraph:
    """
    Create and return the LangGraph workflow for knowledge base processing.
    
    Returns:
        Compiled StateGraph workflow
    """
    workflow = StateGraph(KnowledgeBaseState)
    
    # Add nodes
    workflow.add_node("initialize", initialize_workbook)
    workflow.add_node("process_sheets", process_all_sheets)
    workflow.add_node("upsert_to_qdrant", upsert_facts_to_qdrant)
    
    # Define the flow
    workflow.set_entry_point("initialize")
    workflow.add_edge("initialize", "process_sheets")
    workflow.add_edge("process_sheets", "upsert_to_qdrant")
    workflow.add_edge("upsert_to_qdrant", END)
    
    # Compile the graph
    return workflow.compile()


# Main function to run the workflow
def process_knowledge_base(
    sectors: List[str],
    technologies: List[str],
    workbook_path: str
) -> Dict[str, Any]:
    """
    Process an Excel workbook with sectors and technologies using LangGraph flow.
    
    Args:
        sectors: List of sector names
        technologies: List of technology names
        workbook_path: Path to the Excel workbook file
        
    Returns:
        Dictionary containing:
            - sectors: List of sectors
            - technologies: List of technologies
            - sheet_results: Dictionary mapping sheet names to their data (table_text and batches)
            - errors: List of any errors encountered
    """
    # Validate inputs
    if not sectors or len(sectors) == 0:
        raise ValueError("sectors list must contain at least one sector")
    if not technologies or len(technologies) == 0:
        raise ValueError("technologies list must contain at least one technology")
    
    # Create initial state
    initial_state: KnowledgeBaseState = {
        "sectors": sectors,
        "technologies": technologies,
        "workbook_path": workbook_path,
        "sheet_names": [],
        "sheet_results": {},
        "errors": [],
        "facts_upserted": False
    }
    
    # Create and run the workflow
    workflow = create_knowledge_base_workflow()
    final_state = workflow.invoke(initial_state)
    
    # Return results
    return {
        "sectors": final_state["sectors"],
        "technologies": final_state["technologies"],
        "sheet_results": final_state["sheet_results"],
        "errors": final_state["errors"]
    }


if __name__ == "__main__":
    # Test the knowledge base processing workflow
    import os
    from pathlib import Path
    from datetime import datetime
    
    # Get the project root directory (parent of app directory)
    # This file is in app/services/, so we go up 2 levels to get project root
    current_file = Path(__file__).resolve()
    project_root = current_file.parent.parent.parent
    static_dir = project_root / "app" / "static"
    
    # Example test data - modify these paths and values as needed
    # Option 1: Use file from app/static directory
    test_workbook_path = static_dir / "Alexandria Real Estate (ARE) - Bon Appetit_Agilysys_SquareSpace_Triple Seat_Authorize.net {05. IT Control Evaluation and Implementation Testwork - PBC}.xlsx"  # File from app/static
    # Option 2: Use a file in the current directory or provide absolute path
    # test_workbook_path = "highlighted_example.xlsx"  # File in current working directory
    
    test_sectors = ["Technology", "Finance", "Healthcare"]
    test_technologies = ["AI", "Cloud Computing", "Blockchain"]
    
    # Convert Path to string for compatibility
    test_workbook_path_str = str(test_workbook_path)
    
    # Prepare output file path
    output_dir = project_root / "logs"
    output_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_dir / f"knowledge_base_output_{timestamp}.txt"
    
    # Function to write to both console and file
    output_lines = []
    
    def write_output(text):
        """Write text to both console and output buffer."""
        print(text)
        output_lines.append(text)
    
    # Check if test file exists
    if not os.path.exists(test_workbook_path_str):
        write_output(f"⚠️  Test file not found: {test_workbook_path_str}")
        write_output(f"   Looking in: {static_dir}")
        write_output("   Available files in app/static:")
        if static_dir.exists():
            for file in static_dir.iterdir():
                if file.is_file():
                    write_output(f"     - {file.name}")
        write_output("\n   Please update 'test_workbook_path' with a valid Excel file path")
    else:
        write_output("=" * 60)
        write_output("Testing Knowledge Base Processing Workflow")
        write_output("=" * 60)
        write_output(f"\n📊 Workbook: {test_workbook_path_str}")
        write_output(f"🏢 Sectors: {test_sectors}")
        write_output(f"🔧 Technologies: {test_technologies}")
        write_output("\n🚀 Starting processing...\n")
        
        try:
            # Run the workflow
            results = process_knowledge_base(
                sectors=test_sectors,
                technologies=test_technologies,
                workbook_path=test_workbook_path_str
            )
            
            # Display results
            write_output("✅ Processing completed successfully!\n")
            write_output("-" * 60)
            write_output("RESULTS:")
            write_output("-" * 60)
            write_output(f"\n📋 Sectors: {results['sectors']}")
            write_output(f"🔧 Technologies: {results['technologies']}")
            write_output(f"\n📄 Sheets Processed: {len(results['sheet_results'])}")
            
            # Display sheet results summary
            for sheet_name, sheet_data in results['sheet_results'].items():
                # Handle both old format (string) and new format (dict)
                if isinstance(sheet_data, str):
                    table_text = sheet_data
                    batches = []
                else:
                    table_text = sheet_data.get("table_text", "")
                    batches = sheet_data.get("batches", [])
                
                lines_count = len(table_text.split('\n'))
                chars_count = len(table_text)
                write_output(f"\n  📑 Sheet: {sheet_name}")
                write_output(f"     - Lines: {lines_count}")
                write_output(f"     - Characters: {chars_count:,}")
                write_output(f"     - Preview (first 200 chars): {table_text[:200]}...")
                write_output(f"     - Question Batches: {len(batches)}")
                
                # Display question batch information
                if batches:
                    batch_facts = sheet_data.get("batch_facts", {}) if isinstance(sheet_data, dict) else {}
                    write_output(f"\n     Question Batches:")
                    for i, batch in enumerate(batches, 1):
                        row_nums = batch.get("row_numbers", [])
                        structure = batch.get("structure_info", "")
                        write_output(f"       Batch {i}: Rows {row_nums}")
                        write_output(f"         Structure: {structure}")
                        
                        # Extract and display batch table text
                        if row_nums:
                            batch_table_text = extract_batch_table_text(table_text, row_nums)
                            if batch_table_text:
                                batch_lines = len(batch_table_text.split('\n'))
                                batch_chars = len(batch_table_text)
                                write_output(f"         Batch Table Text: {batch_lines} lines, {batch_chars} chars")
                                write_output(f"         Batch Table Text Preview:")
                                # Show first few lines of batch table text
                                preview_lines = batch_table_text.split('\n')[:10]
                                for preview_line in preview_lines:
                                    write_output(f"           {preview_line}")
                                if len(batch_table_text.split('\n')) > 10:
                                    write_output(f"           ... ({len(batch_table_text.split('\n')) - 10} more lines)")
                        
                        # Display facts count if available
                        if i - 1 in batch_facts:
                            facts = batch_facts[i - 1]
                            write_output(f"         Facts: {len(facts)} paragraphs")
                            if facts:
                                write_output(f"         First fact preview: {facts[0][:100]}...")
            
            # Display errors if any
            if results['errors']:
                write_output(f"\n⚠️  Errors encountered: {len(results['errors'])}")
                for error in results['errors']:
                    write_output(f"   - {error}")
            else:
                write_output("\n✅ No errors encountered")
            
            write_output("\n" + "=" * 60)
            
        except ValueError as e:
            write_output(f"❌ Validation Error: {e}")
        except FileNotFoundError as e:
            write_output(f"❌ File Not Found: {e}")
        except Exception as e:
            write_output(f"❌ Error during processing: {e}")
            import traceback
            error_trace = traceback.format_exc()
            write_output(error_trace)
        
        # Write all output to file
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(output_lines))
            write_output(f"\n💾 Complete output saved to: {output_file}")
        except Exception as e:
            print(f"\n⚠️  Warning: Could not save output to file: {e}")

